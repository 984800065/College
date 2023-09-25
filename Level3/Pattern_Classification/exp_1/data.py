import numpy as np
import openpyxl as op
import os

wb = op.load_workbook('数据.xlsx')
ws = wb.active

train_data = []
for i in range(3):
    train_data.append([])
    col = i * 3 + 2
    for j in range(10):
        train_data[i].append([])

        row = j + 2
        
        for k in range(3):
            real_col = col + k
            train_data[i][j].append(ws.cell(row, real_col).value)

train_data = np.array(train_data)

wb2 = op.load_workbook('测试点.xlsx')
ws2 = wb2.active

test_data = []

for i in range(4):
    row = i + 2
    test_data.append([])
    for j in range(3):
        col = j + 1
        test_data[i].append(ws2.cell(row, col).value)

test_data = np.array(test_data)

# print(train_data)

# print(test_data)